import datetime
import re
import os
import openpyxl
from urllib.request import urlopen

from PIL import Image
from django.http import Http404
from django.shortcuts import render, get_object_or_404
from django.views.generic.base import TemplateView
from transliterate import translit

from django.contrib.redirects.models import Redirect
from menu.models import PriceHouse
from menu.models import MenuCatalog, House  # , tableProductDop, ProductUsluga
# from menu.models import Standart
from material.models import Material, Floor, Bathroom, Bedroom, AdditionalParam
from project_settings.models import ProjectSettings
from .forms import UploadFileForm
# from .models import ImportData, ExportData, ImportRedirectData, ExportRedirectData
from .models import ImportData, ExportData, ExportPriceData, ImportPriceData


DEFAULT_OPTION = u'Все'

IND_STATE_WORK = 1
IND_STATE_READY = 2
IND_STATE_ERROR = 3
IND_STATE_POTENTIAL_ERROR = 4

ID_TYPE_MENU_CATALOG = 7

COUNT_ROW_FOR_INFO_UPDATE = 100
COUNT_ROW_FOR_INFO_UPDATE_CONTROL = 100
# COUNT_ROW_FOR_INFO_UPDATE_FORMAT = 1000

MAX_COUNT_ROW_INFO = 1000

TYPE_ACTION_IMPORT = "import"
TYPE_ACTION_IMPORT_MARKA = "import_marka"
TYPE_ACTION_IMPORT_HARD = "import_hard"
TYPE_ACTION_CONTROL = "control"

DEFAULT_MESSAGE_ERROR = u"Импорт завершен с ошибками"

COLUMN_ID = 0
COLUMN_ORDER = 1
COLUMN_NAME = 2
COLUMN_NAME_FULL = 3
COLUMN_CATEGORY_ID = 4
COLUMN_FLOOR = 5
COLUMN_BEDROOM = 6
COLUMN_BATHROOM = 7
COLUMN_DOP_PARAM = 8
COLUMN_PRICE = 9
COLUMN_IMAGE_1 = 10
COLUMN_IMAGE_2 = 11
COLUMN_IMAGE_3 = 12
COLUMN_IMAGE_4 = 13
COLUMN_IMAGE_5 = 14
COLUMN_IMAGE_6 = 15

COLUMN_IMAGE_1_HTTP = -2
COLUMN_IMAGE_2_HTTP = -1
COLUMN_IMAGE_1_TECH_NAME = -4
COLUMN_IMAGE_2_TECH_NAME = -3

COLUMN_IMAGE_3_HTTP = -6
COLUMN_IMAGE_4_HTTP = -5
COLUMN_IMAGE_3_TECH_NAME = -8
COLUMN_IMAGE_4_TECH_NAME = -7

COLUMN_IMAGE_5_HTTP = -10
COLUMN_IMAGE_6_HTTP = -9
COLUMN_IMAGE_5_TECH_NAME = -12
COLUMN_IMAGE_6_TECH_NAME = -11

COLUMN_COUNT = 16

ROW_MIN_COUNT = 2

CONTROL_CODE = "73aoF6N"

FLAG_IMPORT = 1


def static_admin_url(request):
	return {
		'static_admin_url': '/static/admin_m/',
	}



def import_image(request, image_name, url_image, info=None, product_id=None):
	from admin_m.views import IND_STATE_ERROR
	
	# Vérifier si image_name est vide
	if not image_name:
		if info:
			info.info += u"ОШИБКА: Имя изображения отсутствует для продукта с id=%d\n" % int(product_id)
			info.state_id = IND_STATE_ERROR
			info.save()
		return

	# Extraire uniquement le nom du fichier à partir de l'URL
	if url_image:
		rx = re.compile('[\/]+')
		rezImg = re.split(rx, url_image)
		image_name = rezImg[-1]  # Prendre le dernier segment comme nom de fichier

	# Créer le chemin d'accès pour l'image
	path_image = os.path.join("www/media/uploads/images/load", image_name)

	# Créer le répertoire si nécessaire
	os.makedirs(os.path.dirname(path_image), exist_ok=True)

	try:
		if url_image.startswith('http'):
			resource = urlopen(url_image)
		else:
			resource = urlopen('http://' + url_image)

		# Vérifier si la réponse est 404
		if resource.code == 404:
			raise Exception("Изображение не найдено (404) : %s" % url_image)

		# Écrire le contenu de l'image dans un fichier
		with open(path_image, 'wb') as out:
			out.write(resource.read())

		# Valider le format de l'image en essayant de l'ouvrir avec PIL
		img = Image.open(path_image).convert('RGB')

	except Exception as e:
		if info:
			info.info += u"ОШИБКА при загрузке изображения дома c id=%d: %s\n" % (int(product_id), str(e))
			info.state_id = IND_STATE_ERROR
			info.save()
		return

class AdminMIndexView(TemplateView):
	"""
	Класс для отображения главной страницы
	"""
	template_name = "admin_m/index.html"
	
	def get(self, request):
		return render(request, self.template_name, locals())


class AdminMImportInfoView(TemplateView):
	"""
	Класс для отображения страницы информации об импорте
	"""
	template_name = "admin_m/control/import/import_info.html"
	
	def get(self, request, import_info_slug):
		current_info = get_object_or_404(ImportData, id=import_info_slug)
		current_info_information = current_info.info.replace("\n", "<br/>")
		return render(request, self.template_name, locals())

class AdminMImportPriceInfoView(TemplateView):
	"""
	Класс для отображения страницы информации об импорте
	"""
	template_name = "admin_m/control/import/import_info.html"
	
	def get(self, request, import_info_slug):
		current_info = get_object_or_404(ImportPriceData, id=import_info_slug)
		current_info_information = current_info.info.replace("\n", "<br/>")
		return render(request, self.template_name, locals())

class AdminMImportView(TemplateView):
	"""
	Класс для отображения страницы импорта
	"""
	template_name = "admin_m/control/import/import.html"
	
	def get(self, request):
		items_list = ImportData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		
		return render(request, self.template_name, locals())
	
	def post(self, request, template_name="admin_m/control/import/import_table.html"):
		post_data = request.POST
		email = None
		try:
			email = post_data["email"]
		except KeyError:
			pass
		
		form = UploadFileForm(request.POST, request.FILES)
		file_name = None
		if form.is_valid():
			main_name = translit(request.FILES['file'].name.strip(), "ru", reversed=True)
			file_name = 'import/files/import_' + get_today() + '_' + main_name
			handle_uploaded_file(request.FILES['file'], file_name)
		else:
			raise Http404()
		
		try:
			type_action = post_data["type_action"]
		except KeyError:
			raise Http404()
		
		flag_hard_import = None
		try:
			flag_hard_import = post_data["flag_hard_import"]
		except KeyError:
			pass
		
		action_str = ""
		if type_action == TYPE_ACTION_IMPORT:
			if flag_hard_import:
				action_str = u"Импорт (принудительный)"
				type_action = TYPE_ACTION_IMPORT_HARD
			else:
				action_str = u"Импорт"
		elif type_action == TYPE_ACTION_CONTROL:
			action_str = u"Проверка дубликатов"
		elif type_action == TYPE_ACTION_IMPORT_MARKA:
			action_str = u"Импорт параметров"
		
		info = ImportData()
		info.name = request.FILES['file'].name
		info.action = action_str
		info.user = request.user.get_full_name()
		info.email = email
		info.file = file_name
		info.state_id = IND_STATE_WORK
		info.result = "0 %"
		info.result_percent = 0
		info.info = u""
		info.save()
		
		import threading
		t = threading.Thread(target=import_data, args=[info, request.scheme, type_action])
		t.setDaemon(True)
		t.start()
		# import_data(info, request.scheme, type_action)
		
		items_list = ImportData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		
		return render(request, template_name, locals())


def send_mail_result(email, user, name, link, link_admin, action, result):
	subject = u"Панель управления %s. %s (файл: %s) - %s" % (ProjectSettings.objects.first().name, action, name, result)
	plain_text = u"""Здравствуйте, %s!\n%s (файл: %s) %s.""" % (user, action, name, result)
	html_text = u"""<h1>Здравствуйте, %s!</h1>\n<p>%s файла: %s - %s. </p>\n\n
		<p>Вы можете посмотреть подробнее по <a href='%s'>прямой ссылке</a> или из <a href='%s'>панели управления</a></p>""" % (
		user, action, name, result, link, link_admin)
	
	project_settings = ProjectSettings.objects.all()
	if project_settings:
		project_settings = project_settings[0]
		from email.mime.text import MIMEText
		from email.mime.multipart import MIMEMultipart
		address = project_settings.tech_email
		# Compose message
		msg = MIMEMultipart('alternative')
		msg['From'] = project_settings.tech_email
		msg['To'] = email
		msg['Subject'] = subject
		
		part1 = MIMEText(plain_text, 'plain', 'UTF-8')
		part2 = MIMEText(html_text, 'html', 'UTF-8')
		msg.attach(part1)
		msg.attach(part2)
		
		import smtplib
		s = smtplib.SMTP(project_settings.tech_mail_server, 465)
		s.ehlo()
		s.esmtp_features["auth"] = "LOGIN PLAIN"
		s.debuglevel = 5
		s.login(project_settings.tech_email, project_settings.tech_email_pass)
		s.sendmail(address, email, msg.as_string())
		s.quit()


def get_today():
	return datetime.datetime.today().strftime("%Y-%m-%d %H:%M")


def handle_uploaded_file(f, file_name):
	filename = 'www/media/' + file_name
	with open(filename, 'wb+') as destination:
		for chunk in f.chunks():
			destination.write(chunk)


def import_data(info, scheme, type_action):
	data_load = load_xls(info)
	if info.state_id == IND_STATE_ERROR:
		info.result = DEFAULT_MESSAGE_ERROR
		info.info += DEFAULT_MESSAGE_ERROR + "\n"
		info.save()
		return
	
	data_format, new_floor_list, new_bedroom_list, new_bathroom_list = format_type(info, data_load, type_action)
	
	if type_action == TYPE_ACTION_IMPORT:
		if info.state_id == IND_STATE_WORK:
			import_data_import(info, data_format)
	elif type_action == TYPE_ACTION_IMPORT_HARD:
		if info.state_id == IND_STATE_WORK:
			import_data_import_hard(info, data_format)
	elif type_action == TYPE_ACTION_CONTROL:
		import_data_control(info, data_format)
	elif type_action == TYPE_ACTION_IMPORT_MARKA:
		import_data_import_marka(info, new_floor_list, new_bedroom_list, new_bathroom_list)
	
	if info.state_id == IND_STATE_ERROR:
		info.result = "Завершен с ошибками"
		info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, устраните ошибки и повторите импорт"
		info.save()
	elif info.state_id == IND_STATE_POTENTIAL_ERROR:
		info.result = "Завершен с ошибками"
		info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, если Вы действительно хотите изменить категорию у существующих позиций, установите флаг 'отключить проверку дубликатов' и повторите импорт"
		info.save()
	else:
		info.state_id = IND_STATE_READY
		info.save()
	
	if info.email:
		url_site = scheme + "://" + ProjectSettings.objects.first().site_name
		link = url_site + info.get_absolute_url()
		link_admin = url_site + "/admin_m/import/"
		send_mail_result(info.email, info.user, info.name, link, link_admin, info.action, info.state.name)
		info.info += u""
		info.save()


def load_xls(info):
	file_name = info.file.path
	data_load = []
	wb = openpyxl.load_workbook(filename=file_name, read_only=True, data_only=True)
	ws = wb.active
	n = 0
	if ws.max_row < ROW_MIN_COUNT:
		info.info += u"Неверный формат файла: количество строк меньше минимального - %d\n" % ROW_MIN_COUNT
		info.state_id = IND_STATE_ERROR
	if ws.max_column != COLUMN_COUNT:
		info.info += u"Неверный формат файла: количество столбцов в файле %d, а должно быть %d\n" % (
			ws.max_column, COLUMN_COUNT)
		info.state_id = IND_STATE_ERROR
	for row in ws.rows:
		# if n == 0:
		#     n += 1
		#     continue
		current_row = []
		for cell in row:
			current_row.append(cell.value)
		if current_row and current_row[0] == None:
			break
		data_load.append(current_row)
	return data_load


def get_floor_number(floor):
	key = floor.replace(' ', '')
	return key


def get_bedroom_number(bedroom):
	key = bedroom.replace(' ', '')
	return key

def get_bathroom_number(bathroom):
	key = bathroom.replace(' ', '')
	return key


def normalize_param_name(param_name):
    return param_name.strip().capitalize()

# def prepare_dop_params(dop_params, info, product_id):
# 	additional_params_objects = []
# 	for param_name in dop_params:
# 		normalized_name = normalize_param_name(param_name)  # Normaliser le nom du paramètre
# 		print(f"Processing parameter: {normalized_name}")  # Afficher le paramètre traité
# 		try:
# 			# Chercher l'AdditionalParam par son nom normalisé
# 			additional_param = AdditionalParam.objects.get(name__iexact=normalized_name)  # Utiliser iexact pour ignorer la casse
# 			additional_params_objects.append(additional_param)
# 			print(f"Found parameter: {additional_param.name}")  # Afficher le paramètre trouvé
# 		except AdditionalParam.DoesNotExist:
# 			# Gérer le cas où le paramètre n'existe pas
# 			# info.info += u"ОШИБКА: Доп. параметр '%s' не существует для дома с id=%d\n" % (param_name, product_id)
# 			# info.save()
# 			print(f"Parameter not found: {param_name}")  # Afficher le paramètre non trouvé
# 			continue  # Passer au paramètre suivant
# 	return additional_params_objects

def prepare_dop_params(dop_params, info, product_id):
    additional_params_objects = []
    
    for param_name in dop_params:
        normalized_name = normalize_param_name(param_name)  # Normaliser le nom du paramètre
        print(f"Processing parameter: {normalized_name}")  # Afficher le paramètre traité
        
        # Chercher l'AdditionalParam par son nom normalisé en utilisant filter() et first()
        additional_param = AdditionalParam.objects.filter(name__iexact=normalized_name).first()
        
        if additional_param:
            additional_params_objects.append(additional_param)
            print(f"Found parameter: {additional_param.name}")  # Afficher le paramètre trouvé
        else:
            # Gérer le cas où le paramètre n'existe pas
            print(f"Parameter not found: {normalized_name}")  # Afficher le paramètre non trouvé
            # Vous pouvez également enregistrer un message d'erreur si nécessaire
            # info.info += u"ОШИБКА: Доп. параметр '%s' не существует для дома с id=%d\n" % (param_name, product_id)
            # info.save()
    
    return additional_params_objects


def format_type(info, data_load, type_action):
	data_format = []
	new_floor_list = []
	new_bedroom_list = []
	new_bathroom_list = []
	rez = True
	
	category_list = []
	category_list_map = MenuCatalog.objects.all().only('id').values('id')
	for item in category_list_map:
		category_list.append(item['id'])
	floor_dict = {}
	floor_dict_map = Floor.objects.all().only('id', 'number').values('id', 'number')
	for item in floor_dict_map:
		floor_dict[item['number'].upper().replace(' ', '')] = item['id']
	
	bedroom_dict = {}
	bedroom_dict_map = Bedroom.objects.all().only('id', 'number').values('id', 'number')
	for item in bedroom_dict_map:
		bedroom_dict[item['number'].replace(' ', '').replace(u'\xa0', u'')] = item['id']
	
	bathroom_dict = {}
	bathroom_dict_map = Bathroom.objects.all().only('id', 'number').values('id', 'number')
	for item in bathroom_dict_map:
		bathroom_dict[item['number'].replace(' ', '').replace(u'\xa0', u'')] = item['id']


	info.result = "Подготовка"
	for i in range(1, len(data_load)):

		data_format.append([])
		data_format[i - 1].append(data_load[i][COLUMN_ID])
		data_format[i - 1].append(data_load[i][COLUMN_ORDER])
		data_format[i - 1].append(data_load[i][COLUMN_NAME])
		data_format[i - 1].append(data_load[i][COLUMN_NAME_FULL])

		category_id = ""
		tmp_category_id = ""
		category_id = data_load[i][COLUMN_CATEGORY_ID]
		
		try:
			tmp_category_id = int(category_id)
			category_id = tmp_category_id
		except:
			pass
		
		if not (category_id in category_list):
			info.info += u"%d строка: Отсутствует категория с идентификатором: %d\n" % (i, category_id)
			info.state_id = IND_STATE_ERROR
			info.save()
		
		data_format[i - 1].append(category_id)
		
		if data_load[i][COLUMN_FLOOR]:
			data_load[i][COLUMN_FLOOR] = str(data_load[i][COLUMN_FLOOR])
			floor_number = get_floor_number(data_load[i][COLUMN_FLOOR])
			if floor_number:
				floor_id = ''
				try:
					floor_id = floor_dict[floor_number]
					data_format[i - 1].append(floor_id)
				except:
					info.info += u"%d строка: Отсутствует этажа: %s\n" % (i, data_load[i][COLUMN_FLOOR])
					info.state_id = IND_STATE_ERROR
					if not (data_load[i][COLUMN_FLOOR] in new_floor_list):
						new_floor_list.append(data_load[i][COLUMN_FLOOR])
					data_format[i - 1].append('')
			else:
				data_format[i - 1].append(floor_number)
		else:
			data_format[i - 1].append('')

		if data_load[i][COLUMN_BEDROOM]:
			data_load[i][COLUMN_BEDROOM] = str(data_load[i][COLUMN_BEDROOM])
			bedroom_number = get_bedroom_number(data_load[i][COLUMN_BEDROOM])
			if bedroom_number:
				bedroom_id = ''
				try:
					bedroom_id = bedroom_dict[bedroom_number]
					data_format[i - 1].append(bedroom_id)
				except:
					info.info += u"%d строка: Отсутствует Число спален: %s\n" % (i, data_load[i][COLUMN_BEDROOM])
					info.state_id = IND_STATE_ERROR
					if not (data_load[i][COLUMN_BEDROOM] in new_bedroom_list):
						new_bedroom_list.append(data_load[i][COLUMN_BEDROOM])
					data_format[i - 1].append('')
			else:
				data_format[i - 1].append(bedroom_number)
		else:
			data_format[i - 1].append('')


		if data_load[i][COLUMN_BATHROOM]:
			data_load[i][COLUMN_BATHROOM] = str(data_load[i][COLUMN_BATHROOM])
			bathroom_number = get_bathroom_number(data_load[i][COLUMN_BATHROOM])
			if bathroom_number:
				bathroom_id = ''
				try:
					bathroom_id = bathroom_dict[bathroom_number]
					data_format[i - 1].append(bathroom_id)
				except:
					info.info += u"%d строка: Отсутствует Число санузлов: %s\n" % (i, data_load[i][COLUMN_BATHROOM])
					info.state_id = IND_STATE_ERROR
					if not (data_load[i][COLUMN_BATHROOM] in new_bathroom_list):
						new_bathroom_list.append(data_load[i][COLUMN_BATHROOM])
					data_format[i - 1].append('')
			else:
				data_format[i - 1].append(bathroom_id)
		else:
			data_format[i - 1].append('')


		if data_load[i][COLUMN_DOP_PARAM]:
			# print_all_additional_params()
			dop_params = [param.strip() for param in data_load[i][COLUMN_DOP_PARAM].split(',')]
			print(f"Dop params extracted for row {i}: {dop_params}")  # Afficher les paramètres extraits
			additional_params_objects = prepare_dop_params(dop_params, info, data_load[i][COLUMN_ID])  # Passer l'ID du produit
			data_format[i - 1].append(additional_params_objects)  # Ajouter la liste des objets préparés
		else:
			data_format[i - 1].append([])  # Ajouter une liste vide si aucun paramètre


		
		data_format[i - 1].append(data_load[i][COLUMN_PRICE])
		
		
		image_name_1 = ""
		image_name_1_http = data_load[i][COLUMN_IMAGE_1]
		image_name_2 = ""
		image_name_2_http = data_load[i][COLUMN_IMAGE_2]
		image_name_3 = ""
		image_name_3_http = data_load[i][COLUMN_IMAGE_3]
		image_name_4 = ""
		image_name_4_http = data_load[i][COLUMN_IMAGE_4]
		image_name_5 = ""
		image_name_5_http = data_load[i][COLUMN_IMAGE_5]
		image_name_6 = ""
		image_name_6_http = data_load[i][COLUMN_IMAGE_6]
		
		if data_load[i][COLUMN_IMAGE_1]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_1])
			if rezImg[-1]:
				image_name_1 = rezImg[-1]
			elif rezImg[-2]:
				image_name_1 = rezImg[-2]
		
		if data_load[i][COLUMN_IMAGE_2]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_2])
			if rezImg[-1]:
				image_name_2 = rezImg[-1]
			elif rezImg[-2]:
				image_name_2 = rezImg[-2]

		if data_load[i][COLUMN_IMAGE_3]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_3])
			if rezImg[-1]:
				image_name_3 = rezImg[-1]
			elif rezImg[-2]:
				image_name_3 = rezImg[-2]

		if data_load[i][COLUMN_IMAGE_4]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_4])
			if rezImg[-1]:
				image_name_4 = rezImg[-1]
			elif rezImg[-2]:
				image_name_4 = rezImg[-2]

		if data_load[i][COLUMN_IMAGE_5]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_5])
			if rezImg[-1]:
				image_name_5 = rezImg[-1]
			elif rezImg[-2]:
				image_name_5 = rezImg[-2]

		if data_load[i][COLUMN_IMAGE_6]:
			rx = re.compile('[\/]+')
			rezImg = re.split(rx, data_load[i][COLUMN_IMAGE_6])
			if rezImg[-1]:
				image_name_6 = rezImg[-1]
			elif rezImg[-2]:
				image_name_6 = rezImg[-2]


		image_name_bd = "uploads/images/load/" + image_name_1
		data_format[i - 1].append(image_name_bd)
		image_name_bd = "uploads/images/load/" + image_name_2
		data_format[i - 1].append(image_name_bd)
		image_name_bd = "uploads/images/load/" + image_name_3
		data_format[i - 1].append(image_name_bd)
		image_name_bd = "uploads/images/load/" + image_name_4
		data_format[i - 1].append(image_name_bd)
		image_name_bd = "uploads/images/load/" + image_name_5
		data_format[i - 1].append(image_name_bd)
		image_name_bd = "uploads/images/load/" + image_name_6
		data_format[i - 1].append(image_name_bd)
		
		
		
		data_format[i - 1].append(image_name_1)
		data_format[i - 1].append(image_name_2)
		data_format[i - 1].append(image_name_3)
		data_format[i - 1].append(image_name_4)
		data_format[i - 1].append(image_name_5)
		data_format[i - 1].append(image_name_6)
		data_format[i - 1].append(image_name_1_http)
		data_format[i - 1].append(image_name_2_http)
		data_format[i - 1].append(image_name_3_http)
		data_format[i - 1].append(image_name_4_http)
		data_format[i - 1].append(image_name_5_http)
		data_format[i - 1].append(image_name_6_http)

	
	info.result_percent = 0
	info.save()
	if new_floor_list:
		info.info += u"\nОтсутствуют этажи:\n"
		for item in new_floor_list:
			info.info += u"%s\n" % item
		info.save()
	if new_bedroom_list:
		info.info += u"\nОтсутствуют Число спален:\n"
		for item in new_bedroom_list:
			info.info += u"%s\n" % item
		info.save()
	if new_bathroom_list:
		info.info += u"\nОтсутствуют Число санузлов:\n"
		for item in new_bathroom_list:
			info.info += u"%s\n" % item
		info.save()
	info.info += u"Вы можете загрузить их через 'Импорт параметров'\n"
	info.save()
	return data_format, new_floor_list, new_bedroom_list, new_bathroom_list


def import_data_import_hard(info, data_format):
	list_import_image = []
	count_products = len(data_format)
	for i in range(len(data_format)):
		try:
			base_product = House.objects.get(pk=data_format[i][COLUMN_ID])
			product = base_product
		except House.DoesNotExist:
			product = House()

		info.result_percent += (1. / count_products * 100)
		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
			info.result = ("%.2f" % info.result_percent) + " %"
			info.save()
		
		product.id = data_format[i][COLUMN_ID]
		if data_format[i][1]:
			product.order_number = data_format[i][1]


		if data_format[i][2]:
			product.name = data_format[i][2]		
		else:
			if product.slug:
				product.set_name

		if data_format[i][3]:
			product.name_full = data_format[i][3]
		
		product.catalog_id = data_format[i][4]
		
		if data_format[i][5]:
			product.floor_id = data_format[i][5]
		else:
			product.floor_id = ''
		if data_format[i][6]:
			product.bedroom_id = data_format[i][6]
		else:
			product.bedroom_id = ''
		if data_format[i][7]:
			product.bathroom_id = data_format[i][7]
		else:
			product.bathroom_id = ''
		# Ajouter les dop_param ici
		dop_params_objects = data_format[i][8]  # Récupérer la liste des objets préparés
		if dop_params_objects:  # Vérifier si la liste n'est pas vide
			for additional_param in dop_params_objects:
				if additional_param not in product.dop_param.all():  # Vérifier si déjà associé
					product.dop_param.add(additional_param)  # Associer le paramètre
		else:
			print(f"No new dop_params found for product ID {product.id}. Removing existing dop_params.")  # Message si aucun nouveau paramètre n'est trouvé
			product.dop_param.clear()
			# Afficher les dop_params_objects
			print(f"Dop Params Objects for product ID {product.id}: {[param.name for param in dop_params_objects]}")
		
		product.price = data_format[i][9]
		product.image = data_format[i][10]
		product.image_2 = data_format[i][11]
		product.image_3 = data_format[i][12]
		product.image_4 = data_format[i][13]
		product.image_5 = data_format[i][14]
		product.image_6 = data_format[i][15]
		
		# product.is_home = False
		if not product.is_home:
			product.is_home = False
		if not product.is_hidden:
			product.is_hidden = False
		
		# if product.slug:
		# 	product.set_name()
		# else:
		# 	product.set_slug()

		if not product.slug:
			product.set_slug()

		
		# Контроль на дубли позиций по латинским именам
		try:
			product_tmp = House.objects.only('id').get(slug=product.slug)
			if product_tmp.id != int(product.id):
				info.info += u"ОШИБКА при сохранении продукта - дубликат латинского имени '%s' (текущий id=%d, существующий id=%d)\n" % (
					product.slug, int(product.id), product_tmp.id)
				info.state_id = IND_STATE_ERROR
				info.save()
				break
		except:
			pass
		
		try:
			product.save()
		except Exception as e:
			info.info += u"ОШИБКА при сохранении дома id=%d: %s\n" % (int(product.id), e)
			info.state_id = IND_STATE_ERROR
			info.save()
			break
		
		
		if data_format[i][COLUMN_IMAGE_1_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_1_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				# print("import_image %s %s %s" % (data_format[i][COLUMN_IMAGE_1_TECH_NAME], CONTROL_CODE, data_format[i][COLUMN_IMAGE_1_HTTP]))
				import_image(None, data_format[i][COLUMN_IMAGE_1_TECH_NAME], data_format[i][COLUMN_IMAGE_1_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)
		
		if data_format[i][COLUMN_IMAGE_2_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_2_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				import_image(None, data_format[i][COLUMN_IMAGE_2_TECH_NAME], data_format[i][COLUMN_IMAGE_2_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)

		if data_format[i][COLUMN_IMAGE_3_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_3_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				import_image(None, data_format[i][COLUMN_IMAGE_3_TECH_NAME], data_format[i][COLUMN_IMAGE_3_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)

		if data_format[i][COLUMN_IMAGE_4_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_4_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				import_image(None, data_format[i][COLUMN_IMAGE_4_TECH_NAME], data_format[i][COLUMN_IMAGE_4_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)

		if data_format[i][COLUMN_IMAGE_5_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_5_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				import_image(None, data_format[i][COLUMN_IMAGE_5_TECH_NAME], data_format[i][COLUMN_IMAGE_5_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)

		if data_format[i][COLUMN_IMAGE_6_TECH_NAME]:
			tmp_import_image = data_format[i][COLUMN_IMAGE_6_TECH_NAME]
			if not (tmp_import_image in list_import_image):
				import_image(None, data_format[i][COLUMN_IMAGE_6_TECH_NAME], data_format[i][COLUMN_IMAGE_6_HTTP], info,
							 product.id)
				list_import_image.append(tmp_import_image)

		if info.state_id == IND_STATE_ERROR:
			break
	
	if info.state_id != IND_STATE_ERROR:
		info.result = "100 %"
		info.save()



def import_data_import(info, data_format):
	info.result = "Проверка id категорий - 0 %"
	count_products = len(data_format)
	for i in range(len(data_format)):
		info.result_percent += (1. / count_products * 100)
		# print("control id %.2f   %d/%d" % (info.result_percent, i, count_products))
		if i % COUNT_ROW_FOR_INFO_UPDATE_CONTROL == 0 or i + 1 == count_products:
			info.result = ("Проверка id категорий - %.2f" % info.result_percent) + " %"
			info.save()
		
		try:
			product = House.objects.only('id').get(id=data_format[i][COLUMN_ID])
		except:
			continue
		
		if product.catalog_id != data_format[i][4]:
			info.info += u"%d строка: Внимание! Возможно ошибочно указан id категории %d (сейчас на сайте у данного дома id категории %d) у позиции id=%d\n" % (
				i, data_format[i][4], product.catalog_id, data_format[i][COLUMN_ID])
			if info.state_id == IND_STATE_WORK:
				info.state_id = IND_STATE_POTENTIAL_ERROR
	info.result_percent = 0
	info.save()
	if info.state_id == IND_STATE_WORK:
		import_data_import_hard(info, data_format)


def import_data_control(info, data_format):
	count_products = len(data_format)
	for i in range(len(data_format)):
		info.result_percent += (1. / count_products * 100)
		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
			info.result = ("%.2f" % info.result_percent) + " %"
			info.save()
		
		try:
			product = House.objects.get(id=data_format[i][COLUMN_ID])
		except:
			continue
		
		info.info += u"%d строка: Указанный id уже существует (id - %s) \n" % (i, data_format[i][COLUMN_ID])
		info.state_id = IND_STATE_ERROR
		info.save()


def import_data_import_marka(info, new_floor_list, new_bedroom_list, new_bathroom_list):
	for item in new_floor_list:
		floor = Floor()
		floor.number = item
		floor.set_slug()
		floor.save()
	for item in new_bedroom_list:
		bedroom = Bedroom()
		bedroom.number = item
		bedroom.set_slug()
		bedroom.save()
	for item in new_bathroom_list:
		bathroom = Bathroom()
		bathroom.number = item
		bathroom.set_slug()
		bathroom.save()
	# for item in new_dop_param_list:
	# 	dop_param = AdditionalParam()
	# 	dop_param.name = item
	# 	dop_param.set_slug()
	# 	dop_param.save()
	
	info.result = "100 %"
	info.state_id = IND_STATE_WORK
	info.info += u"\nУспешно загружены"
	info.save()

class AdminMExportView(TemplateView):
	"""
	Класс для отображеия страницы импорта
	"""
	template_name = "admin_m/control/export/export.html"

	def get(self, request):
		default_option = DEFAULT_OPTION
		category_list = MenuCatalog.objects.filter(type_menu_id=ID_TYPE_MENU_CATALOG).only('name',).order_by('name')
		items_list = ExportData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		return render(request, self.template_name, locals())

	def post(self, request, template_name="admin_m/control/export/export_table.html"):
		post_data = request.POST
		name = None
		email = None

		try:
			name = post_data["name"]
			email = post_data["email"]
		except KeyError:
			pass

		info = ExportData()
		info.name = name
		info.user = request.user.get_full_name()
		info.email = email
		info.state_id = IND_STATE_WORK
		info.result = "0 %"
		info.result_percent = 0
		info.save()

		import threading
		t = threading.Thread(target=export_data, args=[info, request.scheme])
		t.setDaemon(True)
		t.start()

		items_list = ExportData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO

		return render(request, template_name, locals())


def export_all_category(info, url_site):
	global COUNT_ROW_FOR_INFO_UPDATE
	COUNT_ROW_FOR_INFO_UPDATE = 100
	category_list = MenuCatalog.objects.filter(type_menu_id=ID_TYPE_MENU_CATALOG).only('id', 'name',)
	count_products = House.objects.all().count()

	d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M")
	name_all_dir = "all_" + d
	os.system("cd www/media/export/;mkdir {0}".format(name_all_dir))
	# info.info = "<p>Начало экспорта всех категорий</p>"
	for item in category_list:
		# info.info += "<p>СТАРТ - Экспорт категории: %s (id = %d) процент: %.2f</p>" % (item.name, item.id, info.result_percent)
		info.save()
		export_data_category(info,
							 url_site,
							 current_category=item,
							 count_products=count_products,
							 name_all_dir=name_all_dir + "/")
		# info.info += "<p>КОНЕЦ - Экспорт категории: %s (id = %d) процент: %.2f</p>" % (item.name, item.id, info.result_percent)
		info.save()
	info.result = "Архивация"
	info.save()
	os.system("cd www/media/export/;rm {0}.tar.gz;tar -zcvf {0}.tar.gz {0}".format(name_all_dir))
	os.system("rm -R www/media/export/{0}".format(name_all_dir))
	info.link = name_all_dir + ".tar.gz"
	info.result = "100 %"
	info.result_percent = 100
	info.save()


def export_data_category(info, url_site, current_category=None, count_products=None, name_all_dir=""):
	flag_all = True
	if not current_category:
		current_category = MenuCatalog.objects.get(name=info.name)
		flag_all = False

	product_list = House.objects.select_related("catalog").filter(catalog_id=current_category.id)
	if not count_products:
		count_products = product_list.count()
	wb = openpyxl.Workbook()
	ws = wb.active  #['Прайс']
	# w = xlwt.Workbook(style_compression=2)
	# ws = w.add_sheet(u'Прайс')
	HEAD_COLOR = 67

	# class Cell(object):
	# 	"""__init__() functions as the class constructor"""
	# 	def __init__(self, data=None, color=int()):
	# 		self.data = data
	# 		self.color = color

	# 	def getData(self):
	# 		return self.data

	# 	def getColor(self):
	# 		return self.color

	i = 1
	# style_date = xlwt.easyxf("pattern: pattern solid; alignment: wrap True, vertical top, horizontal left; border: left thin, right thin, top thin, bottom thin;", num_format_str='DD.MM.YYYY')

	ws.cell(row=i, column=1).value = "ID дома"
	ws.cell(row=i, column=2).value = "Приоритет"
	ws.cell(row=i, column=3).value = "Наименование дома"
	ws.cell(row=i, column=4).value = "Название дома"
	ws.cell(row=i, column=5).value = "Каталог"
	ws.cell(row=i, column=6).value = "Этажность"
	ws.cell(row=i, column=7).value = "Число спален"
	ws.cell(row=i, column=8).value = "Число санузлов"
	ws.cell(row=i, column=9).value = "Доп. параметры"
	ws.cell(row=i, column=10).value = "Базовая цена"
	ws.cell(row=i, column=11).value = "Фото товара 1"
	ws.cell(row=i, column=12).value = "Фото товара 2"
	ws.cell(row=i, column=13).value = "Фото товара 3"
	ws.cell(row=i, column=14).value = "Фото товара 4"
	ws.cell(row=i, column=15).value = "Фото товара 5"
	ws.cell(row=i, column=16).value = "Фото товара 6"

	# ws.write(i, 0, Cell(u"ID товара", HEAD_COLOR).data, style_date)


	# ed_izm = ""
	for item in product_list:
		i = i + 1
		info.result_percent += (1 / count_products * 100)

		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i == count_products:
			if flag_all:
				info.result = ("%.2f" % info.result_percent) + " %"
			else:
				info.result = ("%.0f" % info.result_percent) + " %"
			# info.info += "<p>В процессе - Экспорт категории: %s (id = %d) i=%d процент: %.2f</p>" % (current_category.name, item.id, i, info.result_percent)
			info.save()

		description = ""
		if (item.description):
			description=item.description

		floor=""
		if (item.floor_id):
			floor=item.floor

		bedroom=""
		if (item.bedroom_id):
			bedroom=item.bedroom

		bathroom=""
		if (item.bathroom_id):
			bathroom=item.bathroom

		dop_param=""
		if (item.dop_param):
			dop_param=item.dop_param

		ws.cell(row=i, column=1).value = str(item.id)
		if item.order_number:
			ws.cell(row=i, column=2).value = str(item.order_number)
		else:
			ws.cell(row=i, column=2).value = ''
		# ws.cell(row=i, column=3).value = item.name
		ws.cell(row=i, column=3).value = item.name
		ws.cell(row=i, column=4).value = item.name_full
		ws.cell(row=i, column=5).value = str(item.catalog.id)
		if item.floor:
			ws.cell(row=i, column=6).value = item.floor.number
		else:
			ws.cell(row=i, column=6).value = ""
		if item.bedroom:
			ws.cell(row=i, column=7).value = item.bedroom.number
		else:
			ws.cell(row=i, column=7).value = ""
		if item.bathroom:
			ws.cell(row=i, column=8).value = item.bathroom.number
		else:
			ws.cell(row=i, column=8).value = ""
		if item.dop_param:
			ws.cell(row=i, column=9).value = item.dop_param.name
		else:
			ws.cell(row=i, column=9).value = ""

		ws.cell(row=i, column=10).value = item.price

		if item.image and item.image[-6:] != "/load/":
			ws.cell(row=i, column=11).value = '{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image.replace('/images/load/', '/import_images/')) 
		else:
			ws.cell(row=i, column=11).value = ""
		if item.image_2 and item.image_2[-6:] != "/load/":
			ws.cell(row=i, column=12).value ='{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image_2.replace('/images/load/', '/import_images/'))
		else:
			ws.cell(row=i, column=12).value = ""
		if item.image_3 and item.image_3[-6:] != "/load/":
			ws.cell(row=i, column=13).value ='{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image_3.replace('/images/load/', '/import_images/'))
		else:
			ws.cell(row=i, column=13).value = ""
		if item.image_4 and item.image_4[-6:] != "/load/":
			ws.cell(row=i, column=14).value ='{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image_4.replace('/images/load/', '/import_images/'))
		else:
			ws.cell(row=i, column=14).value = ""
		if item.image_5 and item.image_5[-6:] != "/load/":
			ws.cell(row=i, column=15).value ='{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image_5.replace('/images/load/', '/import_images/'))
		else:
			ws.cell(row=i, column=15).value = ""
		if item.image_6 and item.image_6[-6:] != "/load/":
			ws.cell(row=i, column=16).value ='{}/media/{}'.format(ProjectSettings.objects.first().site_name, item.image_6.replace('/images/load/', '/import_images/'))
		else:
			ws.cell(row=i, column=16).value = ""
		

		# if item.sertificate:
		# 	ws.cell(row=i, column=23).value = str(item.sertificate.id)
		# else:
		# 	ws.cell(row=i, column=23).value = ""
		# ws.cell(row=i, column=25).value = ""

		# products_dop = tableProductDop.objects.filter(product=item)
		# for sub_item in products_dop[:4]:
		#     ws.write(i, tmp_ind, Cell(str(sub_item.product_dop.id)+".0", 1).data, style_date)
		#     tmp_ind += 1
		# tmp_ind = 26
		# while tmp_ind <= 29:
		# 	ws.cell(row=i, column=tmp_ind).value = ""
		# 	tmp_ind += 1

		# usluga_product = ProductUsluga.objects.filter(product=item)
		# tmp_ind = 27
		# for sub_item in usluga_product[:4]:
		#     ws.write(i, tmp_ind, Cell(str(sub_item.usluga.id)+".0", 1).data, style_date)
		#     tmp_ind += 1
		# while tmp_ind <= 33:
		# 	ws.cell(row=i, column=tmp_ind).value = ""
		# 	tmp_ind += 1

	#  создание файла
	d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M")
	id = str(current_category.slug)
	name = id + "_" + d + ".xlsx"
	if not flag_all:
		info.link = name
	wb.save("www/media/export/" + name_all_dir + name)


def export_data(info, scheme):
	url_site = ''
	import time
	if info.name and info.name != DEFAULT_OPTION:
		export_data_category(info, url_site)
	else:
		export_all_category(info, url_site)
	info.state_id = IND_STATE_READY
	info.save()

	url_site = scheme + "://" + ProjectSettings.objects.first().site_name
	if info.email:
		link = url_site + info.get_full_link()
		link_admin = url_site + "/admin_m/export/"
		send_mail_result(info.email, info.user, info.name, link, link_admin, "Экспорт", info.state.name)





class AdminMExportPriceView(TemplateView):
    """
    Класс для отображеия страницы импорта
    """
    template_name = "admin_m/control/export/export_redirect.html"

    def get(self, request):
        default_option = DEFAULT_OPTION
        items_list = ExportPriceData.objects.all()
        items_count_all = items_list.count()
        items_count_show = items_count_all
        if items_count_all > MAX_COUNT_ROW_INFO:
            items_list = items_list[:MAX_COUNT_ROW_INFO]
            items_count_show = MAX_COUNT_ROW_INFO
        return render(request, self.template_name, locals())

    def post(self, request, template_name="admin_m/control/export/export_redirect_table.html"):
        post_data = request.POST
        name = 'Дома/Цена'
        email = None
		
        info = ExportPriceData()
        info.name = name
        info.user = request.user.get_full_name()
        info.email = email
        info.state_id = IND_STATE_WORK
        info.result = "0 %"
        info.result_percent = 0
        info.save()

        import threading
        t = threading.Thread(target=export_data_price, args=[info, request.scheme])
        t.setDaemon(True)
        t.start()

        items_list = ExportPriceData.objects.all()
        items_count_all = items_list.count()
        items_count_show = items_count_all
        if items_count_all > MAX_COUNT_ROW_INFO:
            items_list = items_list[:MAX_COUNT_ROW_INFO]
            items_count_show = MAX_COUNT_ROW_INFO

        return render(request, template_name, locals())

def export_data_price(info, scheme):
    url_site = ''
    import time
    export_all_price(info, url_site)
    info.state_id = IND_STATE_READY
    info.save()

    url_site = scheme + "://" + ProjectSettings.objects.first().site_name

def export_all_price(info, url_site, current_category=None, count_products=None, name_all_dir=""):
	flag_all = True
	product_list = PriceHouse.objects.all()
	if not count_products:
		count_products = product_list.count()
		flag_all = False

	wb = openpyxl.Workbook()
	ws = wb.active
	HEAD_COLOR = 67

	i = 1

	ws.cell(row=i, column=1).value = "ID"
	ws.cell(row=i, column=2).value = "ID дома"
	ws.cell(row=i, column=3).value = "Материал"
	ws.cell(row=i, column=4).value = "Площадь дома"
	ws.cell(row=i, column=5).value = "Цена/Материал"

	for item in product_list:
		i = i + 1
		info.result_percent += (1 / count_products * 100)

		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i == count_products:
			if flag_all:
				info.result = ("%.2f" % info.result_percent) + " %"
			else:
				info.result = ("%.0f" % info.result_percent) + " %"
			info.save()

		ws.cell(row=i, column=1).value = str(item.id)
		ws.cell(row=i, column=2).value = str(item.house.id)
		ws.cell(row=i, column=3).value = str(item.material.name)
		ws.cell(row=i, column=4).value = str(item.surface)
		ws.cell(row=i, column=5).value = str(item.price)
		

    #  создание файла
	d = datetime.datetime.today().strftime("%Y-%m-%d--%H-%M")
	name = "price_" + d + ".xlsx"
	info.link = name
	wb.save("www/media/export/" + name_all_dir + name)

class AdminMImportPriceView(TemplateView):
	"""
	Класс для отображения страницы импорта
	"""
	template_name = "admin_m/control/import/import_redirect.html"
	
	def get(self, request):
		items_list = ImportPriceData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		
		return render(request, self.template_name, locals())
	
	def post(self, request, template_name="admin_m/control/import/import_redirect_table.html"):
		post_data = request.POST
		email = None
		try:
			email = post_data["email"]
		except KeyError:
			pass
		
		form = UploadFileForm(request.POST, request.FILES)
		file_name = None
		if form.is_valid():
			main_name = translit(request.FILES['file'].name.strip(), "ru", reversed=True)
			file_name = 'import/files/import_' + get_today() + '_' + main_name
			handle_uploaded_file(request.FILES['file'], file_name)
		else:
			raise Http404()
		
		try:
			type_action = post_data["type_action"]
		except KeyError:
			raise Http404()
		
		action_str = ""
		if type_action == TYPE_ACTION_IMPORT:
				action_str = u"Импорт"
		
		info = ImportPriceData()
		info.name = request.FILES['file'].name
		info.action = action_str
		info.user = request.user.get_full_name()
		info.email = email
		info.file = file_name
		info.state_id = IND_STATE_WORK
		info.result = "0 %"
		info.result_percent = 0
		info.info = u""
		info.save()
		
		import threading
		t = threading.Thread(target=import_price_data, args=[info, request.scheme, type_action])
		t.setDaemon(True)
		t.start()
		# import_data(info, request.scheme, type_action)
		
		items_list = ImportPriceData.objects.all()
		items_count_all = items_list.count()
		items_count_show = items_count_all
		if items_count_all > MAX_COUNT_ROW_INFO:
			items_list = items_list[:MAX_COUNT_ROW_INFO]
			items_count_show = MAX_COUNT_ROW_INFO
		
		return render(request, template_name, locals())

def load_xlss(info):
	file_name = info.file.path
	data_load = []
	wb = openpyxl.load_workbook(filename=file_name, read_only=True, data_only=True)
	ws = wb.active
	n = 0
	if ws.max_row < ROW_MIN_COUNT:
		info.info += u"Неверный формат файла: количество строк меньше минимального - %d\n" % ROW_MIN_COUNT
		info.state_id = IND_STATE_ERROR
	if ws.max_column != 5:
		info.info += u"Неверный формат файла: количество столбцов в файле %d, а должно быть %d\n" % (
			ws.max_column, 5)
		info.state_id = IND_STATE_ERROR
	for row in ws.rows:
		current_row = []
		for cell in row:
			current_row.append(cell.value)
		if current_row and current_row[0] == None:
			break
		data_load.append(current_row)
	return data_load

def import_price_data(info, scheme, type_action):
	data_load = load_xlss(info)
	if info.state_id == IND_STATE_ERROR:
		info.result = DEFAULT_MESSAGE_ERROR
		info.info += DEFAULT_MESSAGE_ERROR + "\n"
		info.save()
		return
	
	data_format = format_price_type(info, data_load, type_action)
	
	if type_action == TYPE_ACTION_IMPORT:
		if info.state_id == IND_STATE_WORK:
			import_price_data_import(info, data_format)
	
	if info.state_id == IND_STATE_ERROR:
		info.result = "Завершен с ошибками"
		info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, устраните ошибки и повторите импорт"
		info.save()
	elif info.state_id == IND_STATE_POTENTIAL_ERROR:
		info.result = "Завершен с ошибками"
		info.info += u"\nПозиции НЕ ЗАГРУЖЕНЫ, если Вы действительно хотите изменить категорию у существующих позиций, установите флаг 'отключить проверку дубликатов' и повторите импорт"
		info.save()
	else:
		info.state_id = IND_STATE_READY
		info.save()
	
	if info.email:
		url_site = scheme + "://" + ProjectSettings.objects.first().site_name
		link = url_site + info.get_absolute_url()
		link_admin = url_site + "/admin_m/import/"
		send_mail_result(info.email, info.user, info.name, link, link_admin, info.action, info.state.name)
		info.info += u""
		info.save()

def format_price_type(info, data_load, type_action):
	data_format = []
	rez = True
	
	info.result = "Подготовка"
	for i in range(1, len(data_load)):
		
		data_format.append([])
		data_format[i - 1].append(data_load[i][0])
		data_format[i - 1].append(data_load[i][1])
		data_format[i - 1].append(data_load[i][2])
		data_format[i - 1].append(data_load[i][3])
		data_format[i - 1].append(data_load[i][4])

	info.result_percent = 0
	info.save()
	info.save()
	return data_format

def import_price_data_import(info, data_format):
	PriceHouse.objects.all().delete()
	list_import_image = []
	count_products = len(data_format)
	for i in range(len(data_format)):
		product = PriceHouse()
		info.result_percent += (1. / count_products * 100)
		# print("import hard %.2f   %d/%d" % (info.result_percent, i, count_products))
		if i % COUNT_ROW_FOR_INFO_UPDATE == 0 or i + 1 == count_products:
			info.result = ("%.2f" % info.result_percent) + " %"
			info.save()
		
		product.id = data_format[i][0]
		# Récupérer la maison par ID avec vérification
		try:
			product.house = House.objects.get(id=data_format[i][1])  # Utiliser l'ID de la maison
		except House.DoesNotExist:
			info.info += f"Erreur : La maison avec ID {data_format[i][1]} n'existe pas.\n"
			continue  # Passer à l'itération suivante

		# Récupérer le matériau par nom avec vérification
		material_name = data_format[i][2]
		try:
			product.material = Material.objects.get(name=material_name)  # Utiliser le nom du matériau
		except Material.DoesNotExist:
			info.info += f"Erreur : Le matériau '{material_name}' n'existe pas.\n"
			continue  # Passer à l'itération suivante
		
		# Remplir les autres champs
		product.surface = data_format[i][3]
		product.price = data_format[i][4]
		
		try:
			product.save()
		except Exception as e:
			info.info += u"ОШИБКА при сохранении продукта id=%d: %s\n" % (int(product.id), e)
			info.state_id = IND_STATE_ERROR
			info.save()
			break
		
		if info.state_id == IND_STATE_ERROR:
			break
	
	if info.state_id != IND_STATE_ERROR:
		info.result = "100 %"
		info.save()